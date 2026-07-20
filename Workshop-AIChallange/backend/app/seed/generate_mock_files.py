"""Generates placeholder Excel/PDF attachments for Q4/Q5/Q6.

Run standalone (`python -m app.seed.generate_mock_files`) or imported and
called from main.py on startup if the files are missing. These are stand-ins
until the user supplies real workshop attachments (see CLAUDE.md section 5) —
same filenames, same folder layout, safe to overwrite in place later.
"""

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from ..database import UPLOADS_DIR

HEADER_FILL = PatternFill(start_color="16263F", end_color="16263F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def generate_inventory_stock() -> Path:
    out_dir = UPLOADS_DIR / "q4_inventory"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "inventory_stock.xlsx"

    categories = {
        "Screws": ["Hex Head Screw M6x20", "Socket Cap Screw M8x30", "Self-Tapping Screw M4x16",
                   "Machine Screw M5x25", "Wood Screw M6x40"],
        "Bolts": ["Hex Bolt M10x50", "Carriage Bolt M8x60", "U-Bolt M12x80",
                  "Flange Bolt M10x35", "Eye Bolt M8x100"],
        "Nuts": ["Hex Nut M10", "Lock Nut M8", "Wing Nut M6", "T-Nut M8", "Flange Nut M10"],
        "Washers": ["Flat Washer M10", "Spring Washer M8", "Fender Washer M6",
                    "Lock Washer M10", "Square Washer M12"],
    }
    warehouses = ["WH-Selangor", "WH-Penang", "WH-Johor"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = ["SKU", "Product Name", "Category", "Warehouse", "Current Stock",
               "Reorder Point", "Avg Monthly Usage", "Lead Time (days)", "Unit Cost (RM)"]
    ws.append(headers)
    _style_header(ws, len(headers))

    rng = random.Random(42)
    sku_counter = 1000
    for category, products in categories.items():
        for product in products:
            wh = rng.choice(warehouses)
            sku_counter += 1
            avg_usage = rng.randint(80, 600)
            reorder_point = round(avg_usage * rng.uniform(0.8, 1.3))
            # Deliberately mix: below reorder point, borderline, healthy
            bucket = rng.choice(["below", "below", "borderline", "healthy", "healthy"])
            if bucket == "below":
                current_stock = round(reorder_point * rng.uniform(0.2, 0.85))
            elif bucket == "borderline":
                current_stock = round(reorder_point * rng.uniform(0.9, 1.15))
            else:
                current_stock = round(reorder_point * rng.uniform(1.5, 3.0))
            lead_time = rng.choice([7, 10, 14, 21, 30])
            unit_cost = round(rng.uniform(0.15, 8.5), 2)
            ws.append([
                f"BOS-{sku_counter}", product, category, wh,
                current_stock, reorder_point, avg_usage, lead_time, unit_cost,
            ])

    widths = [12, 26, 12, 14, 14, 14, 18, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(path)
    return path


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles["Heading1"].textColor = "#16263F"
    styles["Heading2"].textColor = "#4098CA"
    return styles


def generate_supplier_contract() -> Path:
    out_dir = UPLOADS_DIR / "q5_finance"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "supplier_contract.pdf"

    styles = _pdf_styles()
    doc = SimpleDocTemplate(str(path), pagesize=A4,
                             topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    story = [
        Paragraph("Supply Agreement", styles["Title"]),
        Paragraph("Between Bossard (Malaysia) Sdn Bhd (\"Buyer\") and "
                   "Precision Metal Components Sdn Bhd (\"Supplier\")", styles["Normal"]),
        Spacer(1, 0.6 * cm),

        Paragraph("1. Payment Terms", styles["Heading2"]),
        Paragraph(
            "Buyer shall pay all undisputed invoices within 45 days of the invoice date, "
            "net of a 2% early-payment discount if paid within 10 days. All payments shall "
            "be made in Malaysian Ringgit (MYR) unless otherwise agreed in writing.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("2. Late Delivery Penalty", styles["Heading2"]),
        Paragraph(
            "If Supplier fails to deliver goods within 5 business days of the agreed "
            "delivery date, Supplier shall pay a penalty of 1.5% of the affected order "
            "value per week of delay, capped at 10% of the total order value. Delays "
            "exceeding 6 weeks entitle Buyer to cancel the affected order without penalty.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("3. Termination", styles["Heading2"]),
        Paragraph(
            "Either party may terminate this Agreement with 90 days' written notice. "
            "Buyer may terminate immediately, without notice, in the event of three or "
            "more late-delivery penalty events within any rolling 6-month period, or upon "
            "Supplier's insolvency.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("4. Currency / FX Adjustment", styles["Heading2"]),
        Paragraph(
            "Unit prices are fixed in MYR for the first 12 months. Thereafter, if the "
            "USD/MYR exchange rate moves by more than 8% from the rate on the Effective "
            "Date, either party may request a price renegotiation for raw-material-linked "
            "SKUs, to take effect no earlier than 30 days after request.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("5. Confidentiality", styles["Heading2"]),
        Paragraph(
            "Both parties agree to keep pricing, forecasts, and technical drawings "
            "exchanged under this Agreement confidential for a period of 3 years "
            "following termination.",
            styles["Normal"]),
        Spacer(1, 0.6 * cm),

        Paragraph("Effective Date: 1 January 2026 &nbsp;&nbsp; Term: 24 months",
                   styles["Normal"]),
    ]
    doc.build(story)
    return path


def generate_customer_sales() -> Path:
    out_dir = UPLOADS_DIR / "q6_combined"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "customer_sales.xlsx"

    months = []
    today = date.today().replace(day=1)
    for i in range(5, -1, -1):
        m = (today.month - i - 1) % 12 + 1
        y = today.year + ((today.month - i - 1) // 12)
        months.append(date(y, m, 1).strftime("%b %Y"))

    customers = [
        ("Delta Manufacturing Sdn Bhd", "Johor", "Bolts", "declining"),
        ("Alpine Precision Works", "Penang", "Screws", "declining"),
        ("Cahaya Auto Parts", "Selangor", "Nuts & Washers", "declining"),
        ("Golden Gear Industries", "Perak", "Bolts", "declining"),
        ("Nexa Fabrication Sdn Bhd", "Johor", "Screws", "declining"),
        ("Pinnacle Metal Works", "Selangor", "Bolts", "stable"),
        ("Rimba Engineering", "Sabah", "Nuts & Washers", "stable"),
        ("Sunrise Components", "Penang", "Screws", "stable"),
        ("Titan Structural Sdn Bhd", "Selangor", "Bolts", "growing"),
        ("Vertex Assembly Co", "Johor", "Screws", "growing"),
        ("Ocean Line Marine Supply", "Sabah", "Bolts", "stable"),
        ("Harmoni Plastics & Metal", "Selangor", "Nuts & Washers", "declining"),
        ("Everest Tooling Sdn Bhd", "Penang", "Screws", "growing"),
        ("Bintang Timur Manufacturing", "Perak", "Bolts", "stable"),
        ("Quantum Fastener Solutions", "Selangor", "Nuts & Washers", "growing"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Sales"
    headers = ["Customer Name", "Region", "Product Category"] + months + ["Last Order Date"]
    ws.append(headers)
    _style_header(ws, len(headers))

    rng = random.Random(7)
    for name, region, category, trend in customers:
        base = rng.randint(400, 1200)
        volumes = []
        for i in range(6):
            if trend == "declining":
                factor = 1 - (i * rng.uniform(0.10, 0.16))
            elif trend == "growing":
                factor = 1 + (i * rng.uniform(0.06, 0.12))
            else:
                factor = 1 + rng.uniform(-0.05, 0.05)
            volumes.append(max(20, round(base * factor)))
        if trend == "declining":
            last_order_days_ago = rng.randint(25, 55)
        else:
            last_order_days_ago = rng.randint(1, 15)
        last_order = (date.today() - timedelta(days=last_order_days_ago)).isoformat()
        ws.append([name, region, category, *volumes, last_order])

    widths = [28, 12, 16] + [11] * 6 + [16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(path)
    return path


def generate_market_report() -> Path:
    out_dir = UPLOADS_DIR / "q6_combined"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "market_report.pdf"

    styles = _pdf_styles()
    doc = SimpleDocTemplate(str(path), pagesize=A4,
                             topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    story = [
        Paragraph("Regional Fastener Market Snapshot — Q2 2026", styles["Title"]),
        Spacer(1, 0.4 * cm),

        Paragraph("Competitor Pricing", styles["Heading2"]),
        Paragraph(
            "A regional competitor (IronGrip Fasteners) cut list prices on standard "
            "bolts and screws by 6-9% in the last quarter, aggressively targeting "
            "mid-size manufacturers in Johor and Selangor. Several Bossard accounts in "
            "these regions have reduced order volumes over the same period.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("Demand Shift", styles["Heading2"]),
        Paragraph(
            "Overall demand for nuts &amp; washers has softened slightly due to a "
            "slowdown in the plastics-and-metal fabrication segment, while demand for "
            "screws from precision tooling customers continues to grow, driven by "
            "electronics-sector expansion in Penang.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("Raw Material Costs", styles["Heading2"]),
        Paragraph(
            "Stainless steel raw material costs have risen approximately 4% quarter "
            "over quarter, putting margin pressure on suppliers who compete purely on "
            "price rather than service or lead time.",
            styles["Normal"]),
        Spacer(1, 0.4 * cm),

        Paragraph("Implication for Account Management", styles["Heading2"]),
        Paragraph(
            "Customers who have reduced order volume recently may be splitting spend "
            "with lower-priced competitors rather than leaving the category entirely. "
            "Win-back offers emphasizing reliability, lead time, and bundled service "
            "(rather than matching price cuts directly) are likely to be more "
            "sustainable than a pure discount war.",
            styles["Normal"]),
    ]
    doc.build(story)
    return path


def generate_all() -> list[Path]:
    return [
        generate_inventory_stock(),
        generate_supplier_contract(),
        generate_customer_sales(),
        generate_market_report(),
    ]


if __name__ == "__main__":
    for p in generate_all():
        print(f"generated {p}")
